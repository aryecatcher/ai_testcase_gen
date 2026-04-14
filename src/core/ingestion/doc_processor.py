import json
import re
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
from loguru import logger
from docling.document_converter import DocumentConverter
try:
    from .smart_hierarchical_parser import SmartHierarchicalParser
except ImportError:
    from src.core.ingestion.smart_hierarchical_parser import SmartHierarchicalParser
import openpyxl

class DocProcessor:
    """
    Handles file reading using IBM Docling and other libraries.
    Supports: .docx, .pdf, .html, .pptx, .md, .txt, .xlsx, .json (Swagger/OpenAPI)
    """
    def __init__(self):
        self.converter = DocumentConverter()
        self.hierarchical_parser = SmartHierarchicalParser()

    def read_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Reads a file and returns a list of raw data chunks.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
            
        logger.info(f"Processing file: {file_path}")
        
        ext = path.suffix.lower()
        
        try:
            if ext in ['.xlsx', '.xls']:
                return self._read_excel(path)
            elif ext == '.json':
                return self._read_json(path)
            else:
                # Default to Docling for document formats
                return self._read_docling(path)

        except Exception as e:
            logger.error(f"Processing failed for {file_path}: {e}")
            raise

    def _read_docling(self, path: Path) -> List[Dict[str, Any]]:
        # 1. Use Docling for high-precision parsing
        result = self.converter.convert(path)
        
        # 2. Export to AI-readable formats
        markdown_content = result.document.export_to_markdown()
        doc_json = result.document.export_to_dict()
        
        # 3. Intelligent Semantic Chunking
        # Check for large tables that need special handling
        table_chunks = self._table_aware_chunking(doc_json, path)
        if table_chunks:
            return table_chunks
            
        # Fallback to text chunking if no large tables found
        chunks = self._semantic_chunking(markdown_content)
        
        chunk_data = []
        for i, chunk_text in enumerate(chunks):
            ai_content = self._build_markdown_ai_content(
                markdown_text=chunk_text,
                source_name=path.name,
                chunk_index=i,
                total_chunks=len(chunks),
            )
            chunk_data.append({
                "content": ai_content,
                "metadata": {
                    "raw_content": doc_json,
                    "source": str(path),
                    "chunk_index": i,
                    "total_chunks": len(chunks),
                    "status": "ready_for_ai",
                    "docling_version": "v2",
                    "semantic_format": "markdown"
                }
            })
            
        return chunk_data

    def _build_markdown_ai_content(self, markdown_text: str, source_name: str, chunk_index: int, total_chunks: int) -> str:
        return (
            f"# 文档语义片段\n\n"
            f"> 来源文件: {source_name}\n"
            f"> 解析格式: Markdown\n"
            f"> 当前片段: {chunk_index + 1}/{total_chunks}\n"
            f"> 说明: 以下内容已保留标题层级、列表和表格语义，适合后续 AI 进行结构化理解。\n\n"
            f"## 正文内容\n\n"
            f"{markdown_text}"
        )

    def _table_aware_chunking(self, doc_json: dict, path: Path) -> List[Dict[str, Any]]:
        """
        Special logic for splitting large tables while preserving headers.
        """
        chunks = []
        
        # Recursively find tables in doc_json
        tables = []
        def find_tables_recursive(node):
            if isinstance(node, dict):
                if node.get("type") == "table" and "data" in node:
                    tables.append(node)
                for v in node.values():
                    find_tables_recursive(v)
            elif isinstance(node, list):
                for item in node:
                    find_tables_recursive(item)
                    
        find_tables_recursive(doc_json)
        
        if not tables:
            return []
            
        logger.info(f"Found {len(tables)} tables. Applying smart table splitting.")
        
        for t_idx, table in enumerate(tables):
            grid = table.get("data", {}).get("grid", []) if "grid" in table.get("data", {}) else table.get("data", [])
            if not grid or not isinstance(grid, list): continue
            
            # Heuristic: If table is small (<30 rows), let semantic chunking handle it
            if len(grid) < 30:
                continue
                
            # Smart Splitting for Large Tables
            header = grid[0:1] # Assume row 0 is header
            body = grid[1:]
            
            chunk_size = 20 # Rows per chunk
            overlap = 5 # Overlap rows
            
            for i in range(0, len(body), chunk_size):
                # Slice with overlap logic (start slightly earlier)
                start = max(0, i - overlap) if i > 0 else 0
                end = min(len(body), i + chunk_size)
                
                # Construct chunk: Header + Overlap/Body
                chunk_rows = header + body[start:end]
                
                # Convert back to Markdown table string
                md_table = self._grid_to_markdown(chunk_rows)
                
                # Add Context Header
                context_header = f"# Table Part {i//chunk_size + 1} (Rows {start}-{end})\n"
                context_header += f"> Context: Continuation of Table {t_idx+1} from {path.name}\n\n"
                
                full_content = context_header + md_table
                
                chunks.append({
                    "content": full_content,
                    "metadata": {
                        "raw_content": table,
                        "source": str(path),
                        "table_id": f"TBL-{t_idx}",
                        "is_table_part": True,
                        "part_index": i // chunk_size,
                        "status": "ready_for_ai"
                    }
                })
                
        return chunks

    def _grid_to_markdown(self, grid: List[List[Any]]) -> str:
        """Helper to convert list of lists to MD table"""
        if not grid: return ""
        
        def cell_text(c):
            if isinstance(c, dict): return str(c.get("text", "")).strip().replace("\n", " ")
            return str(c).strip().replace("\n", " ")

        lines = []
        # Header
        headers = [cell_text(c) for c in grid[0]]
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
        
        # Body
        for row in grid[1:]:
            row_txt = [cell_text(c) for c in row]
            lines.append("| " + " | ".join(row_txt) + " |")
            
        return "\n".join(lines)

    def _semantic_chunking(self, text: str, chunk_size: int = 2000, overlap: int = 300) -> List[str]:
        """
        Splits text based on semantic anchors when standard headers (H1/H2) are missing.
        Uses a heuristic approach to find 'Business Entities' or 'Action Verbs'.
        """
        # 1. Cleaning: Remove noise characters
        text = re.sub(r'[ \t]+', ' ', text) # Merge spaces
        text = re.sub(r'\n\s*\n', '\n\n', text) # Merge newlines
        
        # 2. Check if standard headers exist
        has_headers = bool(re.search(r'^#{1,2}\s', text, re.MULTILINE))
        
        if has_headers:
            # Use standard header splitting
            sections = re.split(r'(^#{1,2}\s.*$)', text, flags=re.MULTILINE)
        else:
            # Fallback: Keyword-based splitting (Semantic Anchors)
            # We look for common requirement keywords to start a new section
            # e.g. "功能：", "需求：", "Function:", "Rule:"
            keywords = r'(功能|需求|规则|Constraint|Function|Rule|Scenario|Case)[:：]'
            sections = re.split(f'(^{keywords}.*$)', text, flags=re.MULTILINE)
            
            if len(sections) < 2:
                 # If still no sections, just split by size blindly
                 return [text[i:i+chunk_size] for i in range(0, len(text), chunk_size-overlap)]

        final_chunks = []
        current_chunk = ""
        
        # Re-attach headers/anchors to their content
        if sections[0].strip():
            current_chunk += sections[0]
            
        for i in range(1, len(sections), 2):
            header = sections[i]
            content = sections[i+1] if i+1 < len(sections) else ""
            block = header + content
            
            # If adding this block exceeds chunk size, save current and start new
            if len(current_chunk) + len(block) > chunk_size:
                if current_chunk:
                    final_chunks.append(current_chunk)
                    # Start new chunk with overlap (last N chars of previous)
                    overlap_text = current_chunk[-overlap:] if len(current_chunk) > overlap else current_chunk
                    current_chunk = overlap_text + "\n...[Context Overlap]...\n" + block
                else:
                    current_chunk = block
            else:
                current_chunk += block
                
        if current_chunk:
            final_chunks.append(current_chunk)
            
        return final_chunks if final_chunks else [text]

    def _read_excel(self, path: Path) -> List[Dict[str, Any]]:
        """
        Reads Excel files using Pandas and Calamine, with special handling for hierarchical tables.
        Optimized to avoid redundant file loading.
        """
        chunks = []
        try:
            # 1. Load Workbook once for hierarchy check (need read_only=False for merged_cells)
            wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
            
            # 2. Use Calamine for fast sheet enumeration and standard reading
            # Note: Requires 'python-calamine' package
            try:
                xl = pd.ExcelFile(path, engine='calamine')
            except Exception as e:
                logger.warning(f"Calamine engine failed or not installed: {e}. Falling back to default.")
                xl = pd.ExcelFile(path)

            for sheet_name in xl.sheet_names:
                if sheet_name not in wb.sheetnames:
                    continue
                
                ws = wb[sheet_name]
                
                # Check for hierarchical structure using the already loaded worksheet
                if self._is_hierarchical_ws(ws):
                    logger.info(f"Detected hierarchical table in sheet '{sheet_name}'. Using SmartHierarchicalParser.")
                    try:
                        data = self.hierarchical_parser.parse_excel(str(path), sheet_name=sheet_name)
                        # Use pre-generated markdown if available, otherwise generate it
                        ai_md = data.get("ai_markdown")
                        hierarchy_tree = self.hierarchical_parser.to_markdown_tree(data)
                        records = data.get("records", [])
                        headers = data.get("headers", [])
                        ai_md = self._build_excel_ai_content(
                            sheet_name=sheet_name,
                            source_name=path.name,
                            headers=headers,
                            records=records,
                            row_start=0,
                            row_end=len(records),
                            total_rows=len(records),
                            hierarchy_summary=hierarchy_tree,
                        )
                        
                        chunks.append({
                            "content": ai_md,
                            "metadata": {
                                "source": str(path),
                                "sheet": sheet_name,
                                "type": "hierarchical_table",
                                "structure": data['metadata'],
                                "status": "ready_for_ai",
                                "raw_content": {"type": "excel", "data": {sheet_name: data['records']}},
                                "semantic_format": "excel_json"
                            }
                        })
                        continue # Skip standard processing for this sheet
                    except Exception as e:
                        logger.warning(f"SmartHierarchicalParser failed for sheet '{sheet_name}': {e}. Falling back to standard processing.")
                
                # Fallback / Standard Processing (Optimized with Calamine)
                try:
                    df = pd.read_excel(path, sheet_name=sheet_name, engine='calamine')
                except:
                    df = pd.read_excel(path, sheet_name=sheet_name)
                
                # Remove completely empty rows and columns
                df = df.dropna(how='all', axis=0).dropna(how='all', axis=1)
                
                if df.empty:
                    continue

                # Standardize column names (convert to string, strip whitespace)
                df.columns = df.columns.astype(str).str.strip()
                
                # Forward fill to handle merged cells (common in Excel requirements)
                # This propagates values like "System Name" or "Module" down to all relevant rows
                df = df.ffill()
                
                # Smart NaN handling: 
                # - Numeric columns -> keep NaN (or specific marker if needed)
                # - Object/Text columns -> fill with ""
                for col in df.columns:
                    if df[col].dtype == 'object':
                        df[col] = df[col].fillna("").astype(str).str.strip()
                    # For numeric columns, we might leave NaN or fill with 0 based on context, 
                    # but for markdown generation, keeping them as empty string often looks cleaner
                    # or we can use a placeholder like "N/A"
                    
                # Fill remaining NaNs with empty string for display
                df = df.fillna("")

                # 3. Smart Chunking Strategy
                # Default to 30 rows per chunk with 5 rows overlap
                chunk_size = 30
                overlap = 5
                total_rows = len(df)
                
                # If small sheet, process as single chunk
                if total_rows <= 50:
                    chunk_size = total_rows
                    overlap = 0
                
                logger.info(f"Processing sheet '{sheet_name}' ({total_rows} rows). Strategy: Chunk={chunk_size}, Overlap={overlap}")
                
                for i in range(0, total_rows, chunk_size):
                    # Calculate slice range with overlap
                    start = max(0, i - overlap) if i > 0 else 0
                    end = min(total_rows, i + chunk_size)
                    
                    # Avoid infinite loop if start calculation gets stuck (though logic above is safe)
                    if start >= end: break
                    
                    # Slice DataFrame
                    chunk_df = df.iloc[start:end]
                    
                    records = chunk_df.to_dict(orient='records')
                    ai_content = self._build_excel_ai_content(
                        sheet_name=sheet_name,
                        source_name=path.name,
                        headers=[str(c) for c in df.columns],
                        records=records,
                        row_start=start + 1,
                        row_end=end,
                        total_rows=total_rows,
                    )

                    chunks.append({
                        "content": ai_content,
                        "metadata": {
                            "source": str(path),
                            "sheet": sheet_name,
                            "row_start": start,
                            "row_end": end,
                            "total_rows": total_rows,
                            "is_part": True,
                            "status": "ready_for_ai",
                            "raw_content": {"type": "excel", "data": {sheet_name: records}},
                            "semantic_format": self._excel_preferred_format(chunk_df)
                        }
                    })
                    
                    # If we processed the whole table in one go (small table case), break
                    if chunk_size == total_rows:
                        break
                    
        except Exception as e:
            logger.error(f"Error reading Excel {path}: {e}")
            raise

        return chunks

    def _excel_preferred_format(self, df: pd.DataFrame) -> str:
        if len(df.columns) > 20 or len(df) > 80:
            return "excel_csv"
        return "excel_json"

    def _build_excel_ai_content(
        self,
        sheet_name: str,
        source_name: str,
        headers: List[str],
        records: List[Dict[str, Any]],
        row_start: int,
        row_end: int,
        total_rows: int,
        hierarchy_summary: str = "",
    ) -> str:
        fmt = "JSON" if len(headers) <= 20 and len(records) <= 80 else "CSV"
        lines = [
            "# Excel 结构化数据片段",
            "",
            f"> 来源文件: {source_name}",
            f"> Sheet: {sheet_name}",
            f"> 行范围: {row_start}-{row_end} / {total_rows}",
            f"> 推荐读取格式: {fmt}",
            "> 说明: 以下数据已将表头与每行值绑定；请优先依据字段名理解业务含义，避免错列或错行。",
            "",
            "## 元数据说明",
            f"- 列数: {len(headers)}",
            f"- 列名: {', '.join(headers)}",
        ]
        if hierarchy_summary:
            lines.extend([
                "",
                "## 层级摘要",
                hierarchy_summary,
            ])

        if fmt == "JSON":
            lines.extend([
                "",
                "## JSON Records",
                "```json",
                json.dumps(records, ensure_ascii=False, indent=2),
                "```",
            ])
        else:
            df = pd.DataFrame(records, columns=headers)
            csv_text = df.to_csv(index=False)
            lines.extend([
                "",
                "## CSV Data",
                "```csv",
                csv_text,
                "```",
            ])
        return "\n".join(lines)

    def _is_hierarchical_ws(self, ws) -> bool:
        """
        Check if a worksheet contains hierarchical structure (merged cells).
        Optimized to work with an already loaded worksheet object.
        """
        try:
            # ws is an openpyxl Worksheet object
            # Heuristic: Check number of merged cell ranges
            merged_count = len(ws.merged_cells.ranges)
            
            # Also check if it's a reasonably large table
            if ws.max_row < 3:
                return False
                
            # Threshold: more than 3 merged areas suggests structure
            return merged_count > 3
        except Exception as e:
            logger.warning(f"Could not check hierarchy: {e}")
            return False

    def _dataframe_to_markdown(self, df: pd.DataFrame) -> str:
        """
        Custom DataFrame to Markdown converter for better control over formatting.
        - Handles special characters in cells (like pipes |)
        - Aligns columns
        - Handles empty cells cleanly
        """
        if df.empty:
            return ""

        # Helper to sanitize cell content
        def sanitize(val):
            s = str(val).replace("\n", "<br>").replace("|", "\\|")
            return s if s else " " # Return space for empty cells to keep alignment

        # Header
        headers = [sanitize(c) for c in df.columns]
        header_row = "| " + " | ".join(headers) + " |"
        separator_row = "| " + " | ".join(["---"] * len(headers)) + " |"
        
        lines = [header_row, separator_row]
        
        # Data Rows
        for _, row in df.iterrows():
            row_vals = [sanitize(val) for val in row]
            line = "| " + " | ".join(row_vals) + " |"
            lines.append(line)
            
        return "\n".join(lines)

    def _read_json(self, path: Path) -> List[Dict[str, Any]]:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        # Check if Swagger/OpenAPI
        is_swagger = "swagger" in data or "openapi" in data
        
        content = ""
        if is_swagger:
            content = self._parse_swagger(data)
        else:
            content = json.dumps(data, indent=2, ensure_ascii=False)
            
        return [{
            "content": content,
            "metadata": {
                "raw_content": data,
                "source": str(path),
                "type": "swagger" if is_swagger else "json",
                "status": "ready_for_ai"
            }
        }]

    def _parse_swagger(self, data: dict) -> str:
        """Simple Swagger/OpenAPI flattener"""
        lines = []
        info = data.get("info", {})
        lines.append(f"# API Doc: {info.get('title', 'Unknown')} ({info.get('version', '')})")
        lines.append(f"Description: {info.get('description', '')}\n")
        
        paths = data.get("paths", {})
        for path, methods in paths.items():
            for method, details in methods.items():
                summary = details.get("summary", "")
                desc = details.get("description", "")
                lines.append(f"## {method.upper()} {path}")
                lines.append(f"Summary: {summary}")
                if desc: lines.append(f"Description: {desc}")
                
                # Parameters
                params = details.get("parameters", [])
                if params:
                    lines.append("Parameters:")
                    for p in params:
                        lines.append(f"- {p.get('name')} ({p.get('in')}): {p.get('description', '')} [Required: {p.get('required', False)}]")
                
                lines.append("")
                
        return "\n".join(lines)

